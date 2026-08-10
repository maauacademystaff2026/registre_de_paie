"""Script ponctuel : génère le tableau Excel du barème et des règles de paie.
Pas un module de l'application — à exécuter une fois, pas testé/maintenu."""

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BLEU = "1F497D"
GRIS_CLAIR = "F0F3F8"
BLANC = "FFFFFF"

entete_font = Font(bold=True, color=BLANC, size=11)
entete_fill = PatternFill("solid", fgColor=BLEU)
titre_font = Font(bold=True, color=BLEU, size=14)
sous_titre_font = Font(bold=True, size=11)
bordure = Border(*(Side(style="thin", color="B7C3D4"),) * 4)
alt_fill = PatternFill("solid", fgColor=GRIS_CLAIR)
wrap = Alignment(wrap_text=True, vertical="top")
wrap_centre = Alignment(wrap_text=True, vertical="center", horizontal="center")


def largeurs(feuille, valeurs):
    for lettre, largeur in valeurs.items():
        feuille.column_dimensions[lettre].width = largeur


def entete_tableau(feuille, ligne, colonnes, largeurs_cols):
    for i, texte in enumerate(colonnes):
        c = feuille.cell(row=ligne, column=i + 1, value=texte)
        c.font = entete_font
        c.fill = entete_fill
        c.alignment = wrap_centre
        c.border = bordure
    for i, largeur in enumerate(largeurs_cols):
        feuille.column_dimensions[get_column_letter(i + 1)].width = largeur


def ligne_tableau(feuille, ligne, valeurs, alterne):
    for i, valeur in enumerate(valeurs):
        c = feuille.cell(row=ligne, column=i + 1, value=valeur)
        c.border = bordure
        c.alignment = wrap
        if alterne:
            c.fill = alt_fill


classeur = openpyxl.Workbook()

# --- Feuille 1 : Barème -----------------------------------------------------
f1 = classeur.active
f1.title = "Barème"

f1["A1"] = "Barème de tarification — F CFA par heure"
f1["A1"].font = titre_font
f1.merge_cells("A1:C1")
f1.row_dimensions[1].height = 22

f1["A3"] = "Modifiable dans l'application : écran « Paramètres ». Les valeurs ci-dessous sont celles utilisées par défaut."
f1["A3"].font = Font(italic=True, size=9, color="666666")
f1.merge_cells("A3:C3")

entete_tableau(f1, 5, ["Catégorie", "Niveaux inclus", "Tarif (F CFA / heure)"], [22, 45, 22])
lignes_bareme = [
    ("Langue seule", "Tous niveaux (Anglais, Français ou Espagnol enseigné seul)", 2000),
    ("Primaire", "CP, CE1, CE2, CM1, CM2", 1500),
    ("Collège", "6EME, 5EME, 4EME, 3EME", 2000),
    ("Lycée", "2NDE, 1ERE, TLE", 2500),
    ("BTS", "BTS", 3000),
]
for i, ligne in enumerate(lignes_bareme):
    ligne_tableau(f1, 6 + i, ligne, i % 2 == 1)

f1["A12"] = "Priorité : si la matière est UNE langue seule (pas en combinaison, ex. « Maths/Français » ne compte pas), le tarif Langue seule s'applique quel que soit le niveau de l'élève."
f1["A12"].font = Font(italic=True, size=9.5)
f1.merge_cells("A12:C13")
f1["A12"].alignment = wrap

# --- Feuille 2 : Règles -------------------------------------------------
f2 = classeur.create_sheet("Règles de paie")
f2["A1"] = "Règles de calcul de la paie"
f2["A1"].font = titre_font
f2.merge_cells("A1:B1")
f2.row_dimensions[1].height = 22

entete_tableau(f2, 3, ["Règle", "Détail"], [30, 90])
lignes_regles = [
    ("Qui est payé", "Toute personne apparaissant comme intervenant sur une séance payable — quel que soit son rôle déclaré (Enseignant, Direction, Directeur d'études)."),
    ("Séance payable", "Présentiel, Dispensé, Retard (durée entière payée) et Absent (élève absent, professeur présent) sont TOUS payables. Seule « Absence du prof » (le professeur lui-même absent) n'est pas payable."),
    ("Tarif langue vs niveau", "Matière = UNE langue seule (Anglais, Français ou Espagnol, pas en combinaison) -> tarif langue. Sinon -> tarif du niveau de l'élève (voir feuille « Barème »)."),
    ("Résolution du niveau", "Utilise la colonne « Niveau de classe » en priorité ; si non exploitable, recherche un code de niveau dans la colonne « Classe ». Si toujours introuvable, la séance est listée en « à vérifier », jamais payée à 0 F silencieusement."),
    ("Montant d'une séance", "Durée de la séance (en heures) x tarif horaire déterminé ci-dessus."),
    ("Montant brut d'une personne", "Somme des montants exacts de toutes ses séances payables du mois, arrondi UNE SEULE FOIS à la fin (jamais séance par séance)."),
    ("Versement du 15", "Saisi manuellement, personne par personne, dans l'écran de calcul (0 F CFA par défaut)."),
    ("Net à payer", "Montant brut (arrondi) − Versement du 15."),
    ("Personne exclue", "Case à cocher par personne (écran « Paramètres »). Une personne exclue apparaît avec ses heures réelles pour information, mais son montant est forcé à 0 F."),
    ("Intervenant inconnu", "Une personne qui apparaît dans le journal des appels mais pas dans la Feuille de temps n'est jamais payée automatiquement — elle est signalée dans la zone « à vérifier »."),
]
for i, ligne in enumerate(lignes_regles):
    ligne_tableau(f2, 4 + i, ligne, i % 2 == 1)
    f2.row_dimensions[4 + i].height = 32

classeur.save("../Barème et règles de paie.xlsx")
print("Excel genere")
