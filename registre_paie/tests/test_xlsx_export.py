"""Tests de export/xlsx_export.py (§8.4, §8.3)."""

import openpyxl

import db.repository as repo
from core.agregation import agreger
from core.tarification import BAREME_PAR_DEFAUT
from export.xlsx_export import exporter_historique_annee, exporter_resultats
from tests.conftest import faire_seance


def _resultat_exemple():
    seances = [
        faire_seance(
            enseignant="X",
            eleve_nom="DUPONT",
            eleve_prenom="Jean",
            niveau_de_classe="6EME",
            classe="6EME",
            matiere="Mathématiques",
            duree_minutes=60,
        ),
        faire_seance(
            enseignant="Y",
            eleve_nom="MARTIN",
            eleve_prenom="Alice",
            niveau_de_classe="CP",
            classe="CP",
            matiere="ANGLAIS",
            duree_minutes=90,
        ),
    ]
    return agreger(
        seances=seances,
        personnes_connues={"X", "Y"},
        exclusions={"Y"},
        versements_15={"X": 500},
        bareme=BAREME_PAR_DEFAUT,
    )


def test_export_cree_les_deux_feuilles_attendues(tmp_path):
    chemin = tmp_path / "export.xlsx"
    exporter_resultats(_resultat_exemple(), str(chemin))

    classeur = openpyxl.load_workbook(chemin)
    assert classeur.sheetnames == ["Résultats", "Détail des séances"]


def test_feuille_resultats_contient_les_montants_et_le_total(tmp_path):
    chemin = tmp_path / "export.xlsx"
    exporter_resultats(_resultat_exemple(), str(chemin))

    feuille = openpyxl.load_workbook(chemin)["Résultats"]
    lignes = list(feuille.iter_rows(values_only=True))

    assert lignes[0] == (
        "Nom", "Exclu", "Heures payées", "Montant brut (F CFA)",
        "Versement du 15 (F CFA)", "Net à payer (F CFA)",
    )

    par_nom = {ligne[0]: ligne for ligne in lignes[1:]}
    # openpyxl relit une cellule "" écrite comme None -> c'est le comportement
    # normal d'un round-trip Excel, pas un bug de l'export.
    assert par_nom["X"] == ("X", None, 1.0, 2000, 500, 1500)
    assert par_nom["Y"] == ("Y", "Oui", 1.5, 0, 0, 0)
    assert par_nom["TOTAL"][3] == 2000  # montant brut total : Y exclu ne compte pas


def test_feuille_detail_contient_une_ligne_par_seance(tmp_path):
    chemin = tmp_path / "export.xlsx"
    exporter_resultats(_resultat_exemple(), str(chemin))

    feuille = openpyxl.load_workbook(chemin)["Détail des séances"]
    lignes = list(feuille.iter_rows(values_only=True))

    assert lignes[0][0:5] == ("Enseignant", "Élève", "Date", "Niveau", "Matière")
    corps = lignes[1:]
    assert len(corps) == 2

    ligne_x = next(l for l in corps if l[0] == "X")
    assert ligne_x[1] == "DUPONT Jean"
    assert ligne_x[3] == "6EME"
    assert ligne_x[5] == "01:00"
    assert ligne_x[7] == "Niveau"

    ligne_y = next(l for l in corps if l[0] == "Y")
    assert ligne_y[3] == "—"  # langue seule -> pas de niveau résolu
    assert ligne_y[7] == "Langue"


def test_export_historique_annee_regroupe_les_mois_enregistres(tmp_path):
    conn = repo.connecter(":memory:")
    repo.enregistrer_calcul_mensuel(conn, 2026, 6, _resultat_exemple())
    repo.enregistrer_calcul_mensuel(conn, 2026, 7, _resultat_exemple())
    repo.enregistrer_calcul_mensuel(conn, 2025, 12, _resultat_exemple())

    chemin = tmp_path / "historique.xlsx"
    exporter_historique_annee(conn, 2026, str(chemin))

    feuille = openpyxl.load_workbook(chemin).active
    lignes = list(feuille.iter_rows(values_only=True))

    assert feuille.title == "Historique 2026"
    mois_presents = {ligne[0] for ligne in lignes[1:]}
    assert mois_presents == {6, 7}  # pas 2025-12
