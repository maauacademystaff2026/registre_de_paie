"""Export du mois calculé en .xlsx (§8.4) : feuille "Résultats" + feuille
"Détail des séances", pour qu'un montant puisse être retracé ligne par ligne.

Ne reçoit que des structures déjà calculées par `core/` (ResultatCalculMensuel)
— aucune règle de tarification ici, uniquement de la mise en forme (§11).
"""

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

import db.repository as repo
from core.models import ResultatCalculMensuel
from core.utils import minutes_to_hhmm

_EN_TETES_RESULTATS = (
    "Nom",
    "Exclu",
    "Heures payées",
    "Montant brut (F CFA)",
    "Versement du 15 (F CFA)",
    "Net à payer (F CFA)",
)

_EN_TETES_DETAIL = (
    "Enseignant",
    "Élève",
    "Date",
    "Niveau",
    "Matière",
    "Durée (H)",
    "Tarif appliqué (F CFA/h)",
    "Source du tarif",
    "Personnalisé",
    "Montant (F CFA)",
)

_LABEL_SOURCE_TARIF = {"langue": "Langue", "niveau": "Niveau", "eleve": "Élève (forcé)"}


def _ecrire_feuille_resultats(feuille: Worksheet, resultat: ResultatCalculMensuel) -> None:
    feuille.title = "Résultats"
    feuille.append(_EN_TETES_RESULTATS)

    total_heures = 0.0
    total_brut = 0
    total_versement = 0.0
    total_net = 0.0

    for nom in sorted(resultat.resultats_par_personne):
        r = resultat.resultats_par_personne[nom]
        heures = r.heures_payees_minutes / 60
        feuille.append(
            (nom, "Oui" if r.exclu else "", round(heures, 2), r.montant_brut, r.versement_15, r.net_a_payer)
        )
        total_heures += heures
        total_brut += r.montant_brut
        total_versement += r.versement_15
        total_net += r.net_a_payer

    feuille.append(("TOTAL", "", round(total_heures, 2), total_brut, total_versement, total_net))


def _ecrire_feuille_detail(feuille: Worksheet, resultat: ResultatCalculMensuel) -> None:
    feuille.title = "Détail des séances"
    feuille.append(_EN_TETES_DETAIL)

    for nom in sorted(resultat.resultats_par_personne):
        for d in resultat.resultats_par_personne[nom].details:
            feuille.append(
                (
                    nom,
                    f"{d.seance.eleve_nom} {d.seance.eleve_prenom}".strip(),
                    d.seance.date,
                    d.niveau_resolu or "—",
                    d.seance.matiere,
                    minutes_to_hhmm(d.duree_minutes),
                    d.tarif_horaire,
                    _LABEL_SOURCE_TARIF.get(d.source_tarif, d.source_tarif),
                    "Oui" if d.tarif_exceptionnel else "",
                    d.montant,
                )
            )


def exporter_resultats(resultat: ResultatCalculMensuel, chemin_sortie) -> None:
    """Écrit le fichier .xlsx d'export du mois calculé (§8.4).

    `chemin_sortie` est tout ce qu'accepte `openpyxl.Workbook.save` : un
    chemin (str/Path) ou un flux binaire ouvert en écriture (ex. io.BytesIO
    pour un téléchargement Streamlit sans fichier temporaire sur disque).
    """
    classeur = openpyxl.Workbook()
    _ecrire_feuille_resultats(classeur.active, resultat)
    _ecrire_feuille_detail(classeur.create_sheet(), resultat)
    classeur.save(chemin_sortie)


_EN_TETES_HISTORIQUE = (
    "Mois",
    "Nom",
    "Exclu",
    "Heures payées",
    "Montant brut (F CFA)",
    "Versement du 15 (F CFA)",
    "Net à payer (F CFA)",
)


def exporter_historique_annee(conn, annee: int, chemin_sortie) -> None:
    """§8.3 : export de tous les mois enregistrés d'une année en un fichier
    (un résumé par personne et par mois — pour le détail par séance d'un mois
    précis, utiliser `exporter_resultats` au moment du calcul)."""
    classeur = openpyxl.Workbook()
    feuille = classeur.active
    feuille.title = f"Historique {annee}"
    feuille.append(_EN_TETES_HISTORIQUE)

    for annee_mois, mois, _ in repo.lister_mois(conn):
        if annee_mois != annee:
            continue
        calcul = repo.charger_mois(conn, annee_mois, mois)
        for r in calcul.resultats:
            feuille.append(
                (
                    mois,
                    r.nom,
                    "Oui" if r.exclu else "",
                    round(r.heures_payees_minutes / 60, 2),
                    r.montant_brut,
                    r.versement_15,
                    r.net_a_payer,
                )
            )

    classeur.save(chemin_sortie)
