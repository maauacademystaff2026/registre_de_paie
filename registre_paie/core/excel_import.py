"""Import des deux fichiers Excel sources (§2, §8.1) vers core/models.py.

Seule couche autorisée à connaître le format des fichiers .xlsx (§11, principe
directeur n°1) : core/tarification.py et core/agregation.py ne reçoivent plus
jamais une cellule Excel, uniquement des Personne/Seance déjà propres.

Colonnes validées par en-tête (pas par position) : un fichier réordonné ou
avec des colonnes en plus (observé sur les fichiers réels : `#`, `Action`,
`N° INE`, `Début`, `Fin`, `État`, `Comptabilisé`, `Cours n°`... non utilisées
par le calcul) continue de fonctionner. Une colonne requise manquante ou
renommée lève ErreurImport avec son nom explicite (§8.1 : jamais de plantage
silencieux).
"""

import datetime as _dt

import openpyxl

from core.models import Personne, Seance
from core.utils import hhmm_to_minutes


class ErreurImport(Exception):
    """Schéma de colonnes invalide ou valeur illisible dans un fichier source."""


COLONNES_FEUILLE_DE_TEMPS = ("Nom", "Type d'utilisateur")
COLONNES_GESTION_APPELS = (
    "État élève",
    "Noms",
    "Prénoms",
    "Date",
    "Classe",
    "Section",
    "Niveau de classe",
    "Enseignant/ Utilisateur",
    "Matière",
    "Durée (H)",
)


def _normaliser(valeur) -> str:
    """Les cellules réelles contiennent parfois des retours à la ligne et des
    espaces multiples autour du texte (ex. état élève, noms) : on les réduit
    à une chaîne simple, jamais de traitement cellule à cellule en dehors de
    cette fonction."""
    if valeur is None:
        return ""
    return " ".join(str(valeur).split())


def _entetes(feuille) -> list[str | None]:
    premiere_ligne = next(feuille.iter_rows(min_row=1, max_row=1, values_only=True))
    return [_normaliser(v) or None for v in premiere_ligne]


def _index_colonnes(
    entetes: list[str | None], colonnes_requises: tuple[str, ...], nom_fichier: str
) -> dict[str, int]:
    index = {entete: i for i, entete in enumerate(entetes) if entete is not None}
    manquantes = [c for c in colonnes_requises if c not in index]
    if manquantes:
        trouvees = ", ".join(e for e in entetes if e)
        raise ErreurImport(
            f"{nom_fichier} : colonne(s) manquante(s) ou renommée(s) : "
            f"{', '.join(manquantes)}. Colonnes trouvées dans le fichier : {trouvees}."
        )
    return index


def _convertir_date(valeur, numero_ligne: int) -> str:
    """Retourne une date ISO (YYYY-MM-DD). Gère le format texte DD-MM-YYYY
    documenté (§2.2) et, par prudence, une vraie date Excel si le fichier en
    contient une (observé nulle part à ce jour, mais pas garanti pour tous les mois)."""
    if isinstance(valeur, _dt.datetime):
        return valeur.date().isoformat()
    if isinstance(valeur, _dt.date):
        return valeur.isoformat()

    texte = _normaliser(valeur)
    try:
        jour, mois, annee = texte.split("-")
        return _dt.date(int(annee), int(mois), int(jour)).isoformat()
    except (ValueError, AttributeError) as exc:
        raise ErreurImport(
            f"Gestion des appels, ligne {numero_ligne} : date illisible "
            f"({texte!r}), format attendu DD-MM-YYYY."
        ) from exc


def importer_annuaire(chemin: str) -> list[Personne]:
    """§2.1 : une Personne par nom distinct, avec l'ensemble de ses rôles
    déclarés (une même personne peut avoir plusieurs lignes/rôles)."""
    classeur = openpyxl.load_workbook(chemin, data_only=True)
    feuille = classeur.worksheets[0]
    index = _index_colonnes(_entetes(feuille), COLONNES_FEUILLE_DE_TEMPS, "Feuille de temps")

    roles_par_nom: dict[str, list[str]] = {}
    for row in feuille.iter_rows(min_row=2, values_only=True):
        nom = _normaliser(row[index["Nom"]])
        if not nom:
            continue  # ligne vide en fin de tableau
        roles = roles_par_nom.setdefault(nom, [])
        type_utilisateur = _normaliser(row[index["Type d'utilisateur"]])
        if type_utilisateur and type_utilisateur not in roles:
            roles.append(type_utilisateur)

    return [Personne(nom=nom, roles=tuple(roles)) for nom, roles in roles_par_nom.items()]


def importer_seances(chemin: str) -> list[Seance]:
    """§2.2 : une Seance par ligne du journal des appels. Aucun filtrage
    (payable/non payable, niveau résolu/non résolu) ici — c'est le rôle de
    core/agregation.py et core/tarification.py, sur des Seance déjà propres."""
    classeur = openpyxl.load_workbook(chemin, data_only=True)
    feuille = classeur.worksheets[0]
    index = _index_colonnes(_entetes(feuille), COLONNES_GESTION_APPELS, "Gestion des appels")

    seances = []
    for numero_ligne, row in enumerate(feuille.iter_rows(min_row=2, values_only=True), start=2):
        enseignant = _normaliser(row[index["Enseignant/ Utilisateur"]])
        if not enseignant:
            continue  # ligne vide en fin de tableau

        duree_brute = _normaliser(row[index["Durée (H)"]])
        try:
            duree_minutes = hhmm_to_minutes(duree_brute)
        except ValueError as exc:
            raise ErreurImport(
                f"Gestion des appels, ligne {numero_ligne} : durée illisible ({duree_brute!r})."
            ) from exc

        seances.append(
            Seance(
                enseignant=enseignant,
                eleve_nom=_normaliser(row[index["Noms"]]),
                eleve_prenom=_normaliser(row[index["Prénoms"]]),
                date=_convertir_date(row[index["Date"]], numero_ligne),
                classe=_normaliser(row[index["Classe"]]),
                section=_normaliser(row[index["Section"]]),
                niveau_de_classe=_normaliser(row[index["Niveau de classe"]]),
                matiere=_normaliser(row[index["Matière"]]),
                duree_minutes=duree_minutes,
                etat_eleve=_normaliser(row[index["État élève"]]),
            )
        )
    return seances
